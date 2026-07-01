import pandas as pd
import psycopg

file_path = r"C:\Users\Administrator\Downloads\employees_normalized_1.xlsx"

df = pd.read_excel(file_path)

conn = psycopg.connect(
    host="localhost",
    port=5432,
    dbname="dtc",
    user="dtc_user",
    password="dtc_pass"
)

school_cache = {}

with conn:
    with conn.cursor() as cur:
        for _, row in df.iterrows():
            governorate = str(row["المحافظة"]).strip() if pd.notna(row["المحافظة"]) else ""
            wilaya = str(row["الولاية"]).strip() if pd.notna(row["الولاية"]) else ""
            school_name = str(row["اسم المدرسة"]).strip() if pd.notna(row["اسم المدرسة"]) else "Unknown School"

            school_code = str(row["رمز المدرسة"]).strip() if pd.notna(row["رمز المدرسة"]) else ""
            if not school_code:
                school_code = "UNKNOWN_" + school_name.replace(" ", "_").upper()

            employee_name = str(row["اسم الموظف"]).strip() if pd.notna(row["اسم الموظف"]) else ""
            nationality = str(row["الجنسية"]).strip() if pd.notna(row["الجنسية"]) else ""
            contract_type = str(row["نوع العقد"]).strip() if pd.notna(row["نوع العقد"]) else ""
            job_title = str(row["المسمى الوظيفي"]).strip() if pd.notna(row["المسمى الوظيفي"]) else ""
            subject = str(row["المادة الدراسية"]).strip() if pd.notna(row["المادة الدراسية"]) else ""
            service_years = str(row["فترة الخدمة (بالسنوات)"]).strip() if pd.notna(row["فترة الخدمة (بالسنوات)"]) else ""
            hire_date = str(row["تاريخ التعيين"]).strip() if pd.notna(row["تاريخ التعيين"]) else ""

            if not employee_name:
                continue

            if school_code in school_cache:
                school_id = school_cache[school_code]
            else:
                cur.execute("""
                    INSERT INTO schools (governorate, wilaya, school_code, school_name)
                    VALUES (%s, %s, %s, %s)
                    ON CONFLICT (school_code) DO NOTHING
                """, (governorate, wilaya, school_code, school_name))

                cur.execute("""
                    SELECT school_id
                    FROM schools
                    WHERE school_code = %s
                """, (school_code,))
                school_id = cur.fetchone()[0]
                school_cache[school_code] = school_id

            cur.execute("""
                INSERT INTO employees (
                    school_id,
                    employee_name,
                    nationality,
                    contract_type,
                    job_title,
                    subject,
                    service_years,
                    hire_date
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                school_id,
                employee_name,
                nationality,
                contract_type,
                job_title,
                subject,
                service_years,
                hire_date
            ))

conn.close()
print("Import completed successfully.")

import pandas as pd
import psycopg
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

output_path = r"C:\Users\Administrator\Downloads\employees_with_duplicates_highlighted.xlsx"

conn = psycopg.connect(
    host="localhost",
    port=5432,
    dbname="dtc",
    user="dtc_user",
    password="dtc_pass"
)

query = """
SELECT
    s.governorate AS "المحافظة",
    s.wilaya AS "الولاية",
    s.school_code AS "رمز المدرسة",
    s.school_name AS "اسم المدرسة",
    e.employee_name AS "اسم الموظف",
    e.nationality AS "الجنسية",
    e.contract_type AS "نوع العقد",
    e.job_title AS "المسمى الوظيفي",
    e.subject AS "المادة الدراسية",
    e.service_years AS "فترة الخدمة (بالسنوات)",
    e.hire_date AS "تاريخ التعيين"
FROM employees e
JOIN schools s ON e.school_id = s.school_id
ORDER BY s.school_code, e.employee_name
"""

df = pd.read_sql_query(query, conn)
conn.close()

df["رمز المدرسة"] = df["رمز المدرسة"].fillna("").astype(str).str.strip().str.upper()
df["اسم الموظف"] = df["اسم الموظف"].fillna("").astype(str).str.strip()

# color only repeated rows after the first occurrence
df["is_duplicate"] = df.duplicated(
    subset=["رمز المدرسة", "اسم الموظف"],
    keep="first"
)

df_to_export = df.drop(columns=["is_duplicate"])
df_to_export.to_excel(output_path, index=False)

wb = load_workbook(output_path)
ws = wb.active

yellow_fill = PatternFill(
    fill_type="solid",
    start_color="FFFF00",
    end_color="FFFF00"
)

for excel_row, is_dup in zip(range(2, len(df) + 2), df["is_duplicate"]):
    if is_dup:
        for cell in ws[excel_row]:
            cell.fill = yellow_fill

wb.save(output_path)

print("Export completed:", output_path)
print("Repeated rows colored yellow:", int(df["is_duplicate"].sum()))
